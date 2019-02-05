import openpyxl
from openpyxl.styles import Font

from .base import Exporter


class ExcelExporterMixin(object):
    filename_extension = 'xlsx'
    http_content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    def format_header_cell(self, col, name, cell):
        """
        Gets the prepared cell for a header column. Allows manipulation of cell.
        """
        cell.font = Font(bold=True)

    def format_cell(self, row, field, cell):
        """
        Allows formatting of data cells. By manipulating the cell object.
        """
        pass

    def get_sheet(self, workbook):
        sheet = workbook.worksheets[0]
        sheet.title = self.get_export_name()
        start_row = 1
        return (sheet, start_row)

    def write_header(self, sheet, queryset, row=1):
        for col, name in enumerate(self.get_header(queryset), 1):
            cell = sheet['%s%s' % (openpyxl.utils.get_column_letter(col), row)]
            cell.value = name
            self.format_header_cell(col, name, cell)

    def write(self, queryset, fobj=None):
        """
        Generate XLSX data. If fobj is provided, the content is written to that file
        object. If no fobj is provided, the workbook and fobj is returned.
        """
        workbook = openpyxl.Workbook()
        sheet, start_row = self.get_sheet(workbook)

        if self.header:
            self.write_header(sheet, queryset, start_row)
            column_widths = [len(str(cell.value)) + 2 for cell in sheet[start_row]]
            start_row += 1
        else:
            column_widths = None

        data_fields, data = self.get_data(queryset)

        if column_widths is None:
            column_widths = [len(i) for i in data_fields]

        for i, row in enumerate(data, start_row):
            for j, field in enumerate(data_fields, 1):
                value = self.get_value_repr(field, row.get(field, ''))

                if len(value) + 2 > column_widths[j - 1]:
                    column_widths[j - 1] = len(value) + 2

                cell = sheet['%s%s' % (openpyxl.utils.get_column_letter(j), i)]
                cell.value = value
                self.format_cell(row, field, cell)

        for j, width in enumerate(column_widths, 1):
            sheet.column_dimensions[openpyxl.utils.get_column_letter(j)].width = width

        if fobj:
            workbook.save(fobj)
            return fobj

        return (workbook, fobj)


class ExcelExporter(ExcelExporterMixin, Exporter):
    pass

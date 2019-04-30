from io import BytesIO

import pytest
from django.contrib import admin
from django.contrib.auth import get_user_model
from openpyxl import load_workbook

from django_dataexporter.admin import export_csv_action_factory, export_excel_action_factory
from tests.mockapp.models import DummyModel


@pytest.mark.django_db
class TestExport:
    def setup(self):
        self.modeladmin = admin.ModelAdmin(DummyModel, admin.site)
        DummyModel.objects.create(name='Foo', slug='foo', email='foo@bar.baz')

    def test_export_csv_action_factory(self, rf):
        request = rf.get('/')
        fields = ('name', 'email')
        exporter = export_csv_action_factory(fields=fields)
        response = exporter(self, request=request, queryset=DummyModel.objects.all())

        assert response['Content-Type'] == 'text/csv'
        assert response.content == b'name,email\r\nFoo,foo@bar.baz\r\n'

    def test_export_excel_action_factory(self, rf):
        request = rf.get('/')
        fields = ('name', 'email')
        exporter = export_excel_action_factory(fields=fields)
        response = exporter(self, request=request, queryset=DummyModel.objects.all())

        assert response['Content-Type'] == (
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        workbook = load_workbook(filename=BytesIO(response.content))
        sheet = workbook.worksheets[0]

        assert sheet.cell(1, 1).value == 'name'
        assert sheet.cell(1, 2).value == 'email'
        assert sheet.cell(2, 1).value == 'Foo'
        assert sheet.cell(2, 2).value == 'foo@bar.baz'

    def test_use_action_factory_twice(self, rf, admin_client):
        request = rf.get('/')
        User = get_user_model()
        user_id = admin_client.session['_auth_user_id']
        request.user = User.objects.get(pk=user_id)

        exporter_excel = export_excel_action_factory()
        exporter_csv = export_csv_action_factory()
        self.modeladmin.actions += [exporter_excel, exporter_csv]
        actions = self.modeladmin.get_actions(request)

        assert len(actions) == 3
        assert [*actions] == ['delete_selected', 'excelexporter', 'csvexporter']
